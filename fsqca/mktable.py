# -*- coding: utf-8 -*-
# @Author: aka
# @Date:   2018-07-25 16:07:56
# @Last Modified by:   aka
# @Last Modified time: 2018-07-25 21:22:02
# @Email: tenag_hirmb@hotmail.com

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

filename = 'AD_grade_high.txt'
dest_filename = 'fsqca.xlsx'

wb = Workbook()

alignment = Alignment(horizontal='center', vertical='center')

var = ['ME', 'MC', 'CS', 'MAP', 'OP', 'F', 'POP', 'SOP', 'DIS']


def find_paths(data):
    p = re.compile('\S*\*.*\n')
    path = p.findall(data, 68)
    path = list(map(lambda x: x.strip(), path))
    return path


def find_para(data):
    scoverage = re.search('solution coverage: (\S*)', data).group(1)
    sconsistency = re.search('solution consistency: (\S*)', data).group(1)
    return round(float(scoverage), 2), round(float(sconsistency), 2)


def parse_path(line):
    t = re.match('(\S*)\s*(\S*)\s*(\S*)\s*(\S*)', line)
    path = t.group(1)
    path = path.split('*')
    raw_coverage = t.group(2)
    unique_coverage = t.group(3)
    consistency = t.group(4)
    return path, round(float(raw_coverage), 2), round(float(unique_coverage), 2), round(float(consistency), 2)


def magnify(intpath, parsol):
    plist = find_paths(parsol)
    marks = set()
    for p in plist:
        path, _, _, _ = parse_path(p)
        if set(intpath) > set(path):
            marks = marks | set(path)
    return marks


def draw_path(line, ws, parsol, base):
    path, raw_coverage, unique_coverage, consistency = parse_path(line)
    ws.cell(row=11, column=base, value=raw_coverage)
    ws.cell(row=12, column=base, value=unique_coverage)
    ws.cell(row=13, column=base, value=consistency)
    for e in path:
        symbol = '⨂' if e[0] == '~' else '●'
        pos = var.index(e.strip('~'))
        ws.cell(row=pos + 2, column=base, value=symbol)
    marks = magnify(path, parsol)
    for el in marks:
        pos = var.index(el.strip('~'))
        ws.cell(row=pos + 2, column=base).font = Font(size=20)


def draw_solution(intsol, sname, ws, parsol, base=0):
    paths = find_paths(intsol)
    scoverage, sconsistency = find_para(intsol)
    count = len(paths)
    ws.cell(row=1, column=2 + base, value=sname)
    ws.merge_cells(start_row=1, start_column=2 + base, end_row=1, end_column=1 + count + base)
    ws.cell(row=14, column=2 + base, value=scoverage)
    ws.merge_cells(start_row=14, start_column=2 + base, end_row=14, end_column=1 + count + base)
    ws.cell(row=15, column=2 + base, value=sconsistency)
    ws.merge_cells(start_row=15, start_column=2 + base, end_row=15, end_column=1 + count + base)
    for i, d in enumerate(paths):
        draw_path(d, ws, parsol, i + 2 + base)
    return count


def mk_ws(filename):
    ws = wb.create_sheet(title=filename.split('.txt')[0])

    for i in range(len(var)):
        ws['A' + str(i + 2)] = var[i]

    ws['A11'] = '覆盖率\n(raw coverage)'
    ws['A12'] = '净覆盖率\n(unique coverage)'
    ws['A13'] = '一致率\n(consistency)'
    ws['A14'] = '解的覆盖率\n(solution coverage)'
    ws['A15'] = '解的一致率\n(solution consistency)'

    with open('result/' + filename) as f:
        result = f.read()

    p = re.compile("\n{3,}")
    rlist = p.split(result)
    while '' in rlist:
        rlist.remove('')

    intsol = rlist[2]
    rintsol = rlist[5]
    parsol = rlist[1]
    rparsol = rlist[4]

    count = draw_solution(intsol, '评论有用性', ws, parsol)
    rcount = draw_solution(rintsol, '评论无用性', ws, rparsol, count)

    for c in [ws.cell(row=r, column=c) for r in range(1, 16) for c in range(1, 2 + count + rcount)]:
        c.alignment = alignment


def main():
    files = os.listdir('result')

    for file in files:
        mk_ws(file)

    wb.save(filename=dest_filename)


if __name__ == '__main__':
    main()
